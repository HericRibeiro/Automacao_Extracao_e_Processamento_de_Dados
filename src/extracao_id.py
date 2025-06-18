from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl import Workbook
from config import config
import os
import logging

def extracao_id(driver):
    todos_numeros = []

    for i in range(1, 392):
        url = f"{config['url_id']}{i}"
        driver.get(url)

        try:
            WebDriverWait(driver, 16).until(
                EC.presence_of_element_located((By.XPATH, config['secao_numeros']))
            )
            elementos = driver.find_elements(By.XPATH, config['linhas_numeros'])
            numeros = [e.text.strip() for e in elementos if e.text.strip()]
            
            if numeros:
                todos_numeros.extend(numeros)
                logging.info(f"✅ ID {i} válido.")
            else:
                logging.warning(f"⚠️ ID {i} sem dados.")
        except:
            logging.error(f"❌ ID {i} inválido ou sem seção.")

    if os.path.exists(config['arquivo']):
        wb = load_workbook(config['arquivo'])
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "terminais"
        ws.cell(row=1, column=1).value = "Numeros"

    linha_destino = 2

    for numero_bruto in todos_numeros:
        ws.cell(row=linha_destino, column=1).value = numero_bruto
        linha_destino += 1
        
    wb.save(config['arquivo'])
    logging.info(f"📁 {len(todos_numeros)} salvos com sucesso na planilha.")
    logging.info(f"📊 Total de linhas preenchidas: {linha_destino - 2}")