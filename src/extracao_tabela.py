from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
from config import config
import shutil
import os
import logging

def extracao_Tabela(driver, nome_arquivo="extracao_tabela.xlsx", criar_backup=True):

    try:
        logging.info(f"📊 Iniciando extração para {nome_arquivo}...")
        arquivo_existe = os.path.exists(nome_arquivo)
        
        # Cria backup para evitar perda de dados
        if criar_backup and arquivo_existe:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_backup = f"{nome_arquivo.split('.')[0]}_backup_{timestamp}.xlsx"
            shutil.copy2(nome_arquivo, nome_backup)
            logging.info(f"💾 Backup criado: {nome_backup}")

        if arquivo_existe:
            wb = load_workbook(nome_arquivo)
            ws = wb.active
            linha_inicial = ws.max_row + 1
            
            ultimo_id = 0
            for row in range(ws.max_row, 1, -1):
                cell_value = ws.cell(row=row, column=8).value
                if cell_value and isinstance(cell_value, int):
                    ultimo_id = cell_value
                    break
            
            id_sequencial = ultimo_id + 1
            logging.info(f"🔄 Continuando numeração do ID: {id_sequencial}")

        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Empresas"
            
            cabecalho = ["Cliente", "Motivo", "Qtde de Números", "Data Início", "Data Fim", "Status", "Tabela", "ID"] # Colunas ficticias (Para segurança)
            ws.append(cabecalho)
            linha_inicial = 2
            id_sequencial = 1
            logging.info(f"📄 Novo arquivo criado com cabeçalho")

        try:
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, config['tabela']))
            )

        except Exception as e:
            logging.warning(f"⚠️ Tabela não encontrada: {e}")
            return False
        
        linhas = driver.find_elements(By.XPATH, config['linhas_tabela'])
        logging.info(f"🔍 Extraindo {len(linhas)} linhas da página...")

        linha_atual = linha_inicial
        registros_adicionados = 0
        
        for idx, linha in enumerate(linhas, start=1):
            try:
                colunas = linha.find_elements(By.TAG_NAME, "td")[1:]
                if len(colunas) < 6:
                    continue
                
                dados = [col.text.strip() for col in colunas[:6]]
                linha_completa = dados + [None, id_sequencial]
                
                # Insere dados na planilha
                for col_idx, valor in enumerate(linha_completa, start=1):
                    ws.cell(row=linha_atual, column=col_idx, value=valor)

                logging.info(f"✅ Linha {id_sequencial} adicionada")
                id_sequencial += 1
                linha_atual += 1
                registros_adicionados += 1
                
            except Exception as linhaErro:
                logging.error(f"❌ Erro na linha {idx}: {linhaErro}")

        wb.save(nome_arquivo)
        logging.info(f"🎉 Concluído! {registros_adicionados} registros adicionados ao arquivo {nome_arquivo}")
        return True
        
    except Exception as e:
        logging.error(f"❌ Erro na extração da função extracao_Tabela: {type(e).__name__}: {e}")
        return False